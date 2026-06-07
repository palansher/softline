from openpyxl import load_workbook

# Загрузка файла

book = load_workbook(filename="cars.xlsx")

# Получаем активный лист

my_list = book.active

# Чтение по строкам
# for row in my_list.iter_rows(min_row=1,max_col=2,values_only=True):
#     # print(row)
#     print(row[0],'стоит',row[1])

# Чтение по столбцам
# for row in my_list.iter_cols(min_row=1,max_col=2,values_only=True):
#     # print(row)
#     print(row[0],row[1])